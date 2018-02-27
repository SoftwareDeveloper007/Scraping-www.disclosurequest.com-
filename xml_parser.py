# import xml.etree.ElementTree as ET
# import xml.etree as ET
import xml.etree.cElementTree as ET
import openpyxl, csv

xml = """<XML>
<?xml version="1.0"?>
<edgarSubmission>

    <schemaVersion>X0707</schemaVersion>

    <submissionType>D</submissionType>

    <testOrLive>LIVE</testOrLive>

    <primaryIssuer>
        <cik>0001660549</cik>
        <entityName>Advantagene, Inc.</entityName>
        <issuerAddress>
            <street1>440 LEXINGTON STREET</street1>
            <city>AUBURNDALE</city>
            <stateOrCountry>MA</stateOrCountry>
            <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
            <zipCode>02466</zipCode>
        </issuerAddress>
        <issuerPhoneNumber>617-916-5445</issuerPhoneNumber>
        <jurisdictionOfInc>DELAWARE</jurisdictionOfInc>
        <issuerPreviousNameList>
            <value>None</value>
        </issuerPreviousNameList>
        <edgarPreviousNameList>
            <value>None</value>
        </edgarPreviousNameList>
        <entityType>Corporation</entityType>
        <yearOfInc>
            <overFiveYears>true</overFiveYears>
        </yearOfInc>
    </primaryIssuer>

    <relatedPersonsList>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Estuardo</firstName>
                <lastName>Aguilar-Cordova</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Laura</firstName>
                <lastName>Aguilar</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Stephen</firstName>
                <middleName>C.</middleName>
                <lastName>Rocamboli</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Mitchell</firstName>
                <lastName>Finer</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Udi</firstName>
                <lastName>Merav</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Fred</firstName>
                <lastName>Mermelstein</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Glenn</firstName>
                <lastName>Mattes</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Alan</firstName>
                <lastName>Smith</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>c/o 440 Lexington Street</street1>
                <city>Auburndale</city>
                <stateOrCountry>MA</stateOrCountry>
                <stateOrCountryDescription>MASSACHUSETTS</stateOrCountryDescription>
                <zipCode>02466</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Director</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification></relationshipClarification>
        </relatedPersonInfo>
    </relatedPersonsList>

    <offeringData>
        <industryGroup>
            <industryGroupType>Biotechnology</industryGroupType>
        </industryGroup>
        <issuerSize>
            <revenueRange>Decline to Disclose</revenueRange>
        </issuerSize>
        <federalExemptionsExclusions>
            <item>06b</item>
        </federalExemptionsExclusions>
        <typeOfFiling>
            <newOrAmendment>
                <isAmendment>false</isAmendment>
            </newOrAmendment>
            <dateOfFirstSale>
                <value>2015-12-04</value>
            </dateOfFirstSale>
        </typeOfFiling>
        <durationOfOffering>
            <moreThanOneYear>false</moreThanOneYear>
        </durationOfOffering>
        <typesOfSecuritiesOffered>
            <isDebtType>true</isDebtType>
            <isOptionToAcquireType>true</isOptionToAcquireType>
        </typesOfSecuritiesOffered>
        <businessCombinationTransaction>
            <isBusinessCombinationTransaction>false</isBusinessCombinationTransaction>
            <clarificationOfResponse></clarificationOfResponse>
        </businessCombinationTransaction>
        <minimumInvestmentAccepted>0</minimumInvestmentAccepted>
        <salesCompensationList></salesCompensationList>
        <offeringSalesAmounts>
            <totalOfferingAmount>2500000</totalOfferingAmount>
            <totalAmountSold>2200000</totalAmountSold>
            <totalRemaining>300000</totalRemaining>
            <clarificationOfResponse></clarificationOfResponse>
        </offeringSalesAmounts>
        <investors>
            <hasNonAccreditedInvestors>false</hasNonAccreditedInvestors>
            <totalNumberAlreadyInvested>15</totalNumberAlreadyInvested>
        </investors>
        <salesCommissionsFindersFees>
            <salesCommissions>
                <dollarAmount>0</dollarAmount>
            </salesCommissions>
            <findersFees>
                <dollarAmount>0</dollarAmount>
            </findersFees>
            <clarificationOfResponse></clarificationOfResponse>
        </salesCommissionsFindersFees>
        <useOfProceeds>
            <grossProceedsUsed>
                <dollarAmount>0</dollarAmount>
            </grossProceedsUsed>
            <clarificationOfResponse>Other than the payment of salaries and other compensation and benefits, no officer, director or promoter will receive any payments from the proceeds of this offering.</clarificationOfResponse>
        </useOfProceeds>
        <signatureBlock>
            <authorizedRepresentative>false</authorizedRepresentative>
            <signature>
                <issuerName>Advantagene, Inc.</issuerName>
                <signatureName>Estuardo Aguilar-Cordova</signatureName>
                <nameOfSigner>Estuardo Aguilar-Cordova</nameOfSigner>
                <signatureTitle>Chief Executive Officer</signatureTitle>
                <signatureDate>2015-12-21</signatureDate>
            </signature>
        </signatureBlock>
    </offeringData>
</edgarSubmission>
</XML>"""

# def print_path_of_elems(elem, elem_path=""):
#     for child in elem:
#         if not child.getchildren() and child.text:
#             # leaf node with text => print
#             print("%s/%s, %s" % (elem_path, child.tag, child.text))
#         else:
#             # node with child elements => recurse
#             print_path_of_elems(child, "%s/%s" % (elem_path, child.tag))
#
#
# root = ET.fromstring(xml)
# print_path_of_elems(root, root.tag)

class xmlParser():
    def __init__(self, elem, elem_path=""):
        # self.result = []
        self.result = {}
        self.keys = []
        self.parse(elem, elem_path)

    def parse(self, elem, elem_path=""):
        for child in elem:
            if not child.getchildren() and child.text:
                path = "{}/{}".format(elem_path, child.tag)
                path = path.replace("XML/edgarSubmission/", "")
                path_sp = path.split("/")
                path_sp = [convertAbbr(elm) for elm in path_sp]
                path_abbr = "_".join(path_sp)
                value = child.text

                if self.keys.count(path_abbr) == 0:
                    # self.result.append([path, path_abbr, value])
                    self.result[path_abbr] = value
                else:
                    path_abbr_rev = path_abbr + "_{}".format(self.keys.count(path_abbr))
                    # self.result.append([path, path_abbr_rev, value])
                    self.result[path_abbr_rev] = value
                self.keys.append(path_abbr)
            else:
                self.parse(child, "{}/{}".format(elem_path, child.tag))

def convertAbbr(str):
    result = [str[0]] + [l for l in str if l.isupper() or l.isdigit()]
    return "".join(result)

if __name__ == '__main__':
    xml = """<XML>
<?xml version="1.0"?>
<edgarSubmission>

    <schemaVersion>X0707</schemaVersion>

    <submissionType>D</submissionType>

    <testOrLive>LIVE</testOrLive>

    <primaryIssuer>
        <cik>0001653782</cik>
        <entityName>American Industrial Partners Capital Fund VI, L.P.</entityName>
        <issuerAddress>
            <street1>330 MADISON AVENUE</street1>
            <street2>28TH FLOOR</street2>
            <city>NEW YORK</city>
            <stateOrCountry>NY</stateOrCountry>
            <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
            <zipCode>10017</zipCode>
        </issuerAddress>
        <issuerPhoneNumber>(212) 627-2360</issuerPhoneNumber>
        <jurisdictionOfInc>DELAWARE</jurisdictionOfInc>
        <issuerPreviousNameList>
            <value>None</value>
        </issuerPreviousNameList>
        <edgarPreviousNameList>
            <value>None</value>
        </edgarPreviousNameList>
        <entityType>Limited Partnership</entityType>
        <yearOfInc>
            <withinFiveYears>true</withinFiveYears>
            <value>2015</value>
        </yearOfInc>
    </primaryIssuer>

    <relatedPersonsList>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>N/A</firstName>
                <lastName>AIPCF VI, LLC</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>330 Madison Avenue</street1>
                <street2>28th Floor</street2>
                <city>New York</city>
                <stateOrCountry>NY</stateOrCountry>
                <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                <zipCode>10017</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Promoter</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification>The General Partner of the Issuer.</relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Kim</firstName>
                <lastName>Marvin</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>330 Madison Avenue</street1>
                <street2>28th Floor</street2>
                <city>New York</city>
                <stateOrCountry>NY</stateOrCountry>
                <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                <zipCode>10017</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification>Senior Managing Member of the General Partner.</relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>John</firstName>
                <lastName>Becker</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>330 Madison Avenue</street1>
                <street2>28th Floor</street2>
                <city>New York</city>
                <stateOrCountry>NY</stateOrCountry>
                <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                <zipCode>10017</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification>Senior Managing Member of the General Partner.</relationshipClarification>
        </relatedPersonInfo>
        <relatedPersonInfo>
            <relatedPersonName>
                <firstName>Dino</firstName>
                <lastName>Cusumano</lastName>
            </relatedPersonName>
            <relatedPersonAddress>
                <street1>330 Madison Avenue</street1>
                <street2>28th Floor</street2>
                <city>New York</city>
                <stateOrCountry>NY</stateOrCountry>
                <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                <zipCode>10017</zipCode>
            </relatedPersonAddress>
            <relatedPersonRelationshipList>
                <relationship>Executive Officer</relationship>
            </relatedPersonRelationshipList>
            <relationshipClarification>Senior Managing Member of the General Partner.</relationshipClarification>
        </relatedPersonInfo>
    </relatedPersonsList>

    <offeringData>
        <industryGroup>
            <industryGroupType>Pooled Investment Fund</industryGroupType>
            <investmentFundInfo>
                <investmentFundType>Private Equity Fund</investmentFundType>
                <is40Act>false</is40Act>
            </investmentFundInfo>
        </industryGroup>
        <issuerSize>
            <revenueRange>Decline to Disclose</revenueRange>
        </issuerSize>
        <federalExemptionsExclusions>
            <item>06b</item>
            <item>3C</item>
            <item>3C.7</item>
        </federalExemptionsExclusions>
        <typeOfFiling>
            <newOrAmendment>
                <isAmendment>false</isAmendment>
            </newOrAmendment>
            <dateOfFirstSale>
                <value>2015-09-30</value>
            </dateOfFirstSale>
        </typeOfFiling>
        <durationOfOffering>
            <moreThanOneYear>false</moreThanOneYear>
        </durationOfOffering>
        <typesOfSecuritiesOffered>
            <isPooledInvestmentFundType>true</isPooledInvestmentFundType>
        </typesOfSecuritiesOffered>
        <businessCombinationTransaction>
            <isBusinessCombinationTransaction>false</isBusinessCombinationTransaction>
            <clarificationOfResponse></clarificationOfResponse>
        </businessCombinationTransaction>
        <minimumInvestmentAccepted>0</minimumInvestmentAccepted>
        <salesCompensationList>
            <recipient>
                <recipientName>Acalyx Advisors, Inc.</recipientName>
                <recipientCRDNumber>174667</recipientCRDNumber>
                <associatedBDName>Acalyx Advisors, Inc.</associatedBDName>
                <associatedBDCRDNumber>174667</associatedBDCRDNumber>
                <recipientAddress>
                    <street1>330 Madison Avenue</street1>
                    <street2>6th Floor</street2>
                    <city>New York</city>
                    <stateOrCountry>NY</stateOrCountry>
                    <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                    <zipCode>10017</zipCode>
                </recipientAddress>
                <statesOfSolicitationList>
                    <state>AK</state>
                    <description>ALASKA</description>
                    <state>CA</state>
                    <description>CALIFORNIA</description>
                    <state>FL</state>
                    <description>FLORIDA</description>
                    <state>IL</state>
                    <description>ILLINOIS</description>
                    <state>IN</state>
                    <description>INDIANA</description>
                    <state>MA</state>
                    <description>MASSACHUSETTS</description>
                    <state>MO</state>
                    <description>MISSOURI</description>
                    <state>NH</state>
                    <description>NEW HAMPSHIRE</description>
                    <state>NJ</state>
                    <description>NEW JERSEY</description>
                    <state>NY</state>
                    <description>NEW YORK</description>
                    <state>PA</state>
                    <description>PENNSYLVANIA</description>
                    <state>TX</state>
                    <description>TEXAS</description>
                    <state>VA</state>
                    <description>VIRGINIA</description>
                </statesOfSolicitationList>
                <foreignSolicitation>false</foreignSolicitation>
            </recipient>
            <recipient>
                <recipientName>MVision Private Equity Advisers USA LLC</recipientName>
                <recipientCRDNumber>122242</recipientCRDNumber>
                <associatedBDName>MVision Private Equity Advisers USA LLC</associatedBDName>
                <associatedBDCRDNumber>122242</associatedBDCRDNumber>
                <recipientAddress>
                    <street1>41 Madison Avenue</street1>
                    <street2>40th Floor</street2>
                    <city>New York</city>
                    <stateOrCountry>NY</stateOrCountry>
                    <stateOrCountryDescription>NEW YORK</stateOrCountryDescription>
                    <zipCode>10010</zipCode>
                </recipientAddress>
                <statesOfSolicitationList></statesOfSolicitationList>
                <foreignSolicitation>true</foreignSolicitation>
            </recipient>
        </salesCompensationList>
        <offeringSalesAmounts>
            <totalOfferingAmount>Indefinite</totalOfferingAmount>
            <totalAmountSold>1800000000</totalAmountSold>
            <totalRemaining>Indefinite</totalRemaining>
            <clarificationOfResponse></clarificationOfResponse>
        </offeringSalesAmounts>
        <investors>
            <hasNonAccreditedInvestors>false</hasNonAccreditedInvestors>
            <totalNumberAlreadyInvested>79</totalNumberAlreadyInvested>
        </investors>
        <salesCommissionsFindersFees>
            <salesCommissions>
                <dollarAmount>6000000</dollarAmount>
            </salesCommissions>
            <findersFees>
                <dollarAmount>0</dollarAmount>
            </findersFees>
            <clarificationOfResponse></clarificationOfResponse>
        </salesCommissionsFindersFees>
        <useOfProceeds>
            <grossProceedsUsed>
                <dollarAmount>0</dollarAmount>
            </grossProceedsUsed>
            <clarificationOfResponse></clarificationOfResponse>
        </useOfProceeds>
        <signatureBlock>
            <authorizedRepresentative>false</authorizedRepresentative>
            <signature>
                <issuerName>American Industrial Partners Capital Fund VI, L.P.</issuerName>
                <signatureName>/s/ Kim Marvin</signatureName>
                <nameOfSigner>Kim Marvin</nameOfSigner>
                <signatureTitle>Senior Managing Member of the General Partner</signatureTitle>
                <signatureDate>2015-10-14</signatureDate>
            </signature>
        </signatureBlock>
    </offeringData>
</edgarSubmission>
</XML>"""
    xml = xml.replace('<?xml version="1.0"?>', "")
    root = ET.fromstring(xml)
    app = xmlParser(root, root.tag)
    result = app.result
    keys = result.keys()
    keys = list(keys)
    keys = ["'{}'".format(key) for key in keys]
    keys_clause = ", ".join(keys)
    print(keys_clause)

    # wb = openpyxl.Workbook()
    # ws = wb.active
    #
    # for i, (path, path_abbr, value) in enumerate(result):
    #     print(path_abbr, "\t\t\t\t", value)
    #     ws.cell(row=i+2, column=1).value = path
    #     ws.cell(row=i+2, column=2).value = path_abbr
    #     ws.cell(row=i+2, column=3).value = value
    #
    # wb.save("Header.xlsx")
    total_csv_name = "total1.csv"
    total_file = open(total_csv_name, 'w', encoding='utf-8', newline='')
    writer = csv.writer(total_file)